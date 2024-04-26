from selenium.webdriver.common.by import By
import time
import openpyxl


class JustJoinItOfferPage:

    def __init__(self, driver):
        self.driver = driver
        self.job_name_xpath = "//span[contains(@class,'link css-kp67em')]"
        self.tech_stack_class = "css-cjymd2"
        self.company_name_class = "css-mbkv7r"

    def jjitOpenOfferDetail(self, link_list):
        for offer_link in link_list:
            offer_nr = link_list.index(offer_link)
            self.driver.get(offer_link)
            time.sleep(5)
            company_name = self.driver.find_element(By.CLASS_NAME, self.company_name_class).text
            job_name = self.driver.find_element(By.XPATH, self.job_name_xpath).text
            offer_tech_stack_obj = self.driver.find_elements(By.CLASS_NAME, self.tech_stack_class)
            offer_tech_stack = []
            for i in offer_tech_stack_obj: offer_tech_stack.append(i.text)
            name = "Oferty.xlsx"
            exel = openpyxl.load_workbook(name)
            sheet = exel["Oferty"]
            first_empty_row = sheet.max_row + 1
            sheet.cell(row=first_empty_row, column=1, value=offer_nr)
            sheet.cell(row=first_empty_row, column=2, value=company_name)
            sheet.cell(row=first_empty_row, column=3, value=job_name)
            sheet.cell(row=first_empty_row, column=4, value=str(offer_tech_stack))
            sheet.cell(row=first_empty_row, column=5, value=offer_link)
            exel.save(name)
